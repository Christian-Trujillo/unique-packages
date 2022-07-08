import os
from googleapiclient.discovery import build
from google.oauth2 import service_account
import json
import openpyxl as xl


bullet_path = os.getcwd() + r'\sku list & template\Full SKU List Descriptions _ Bullets.xlsx'
bullet_wkbk = xl.load_workbook(bullet_path)
bullet= bullet_wkbk.active

bullet_len = bullet.max_row+1


def create_description_list():
    desc_dict ={}
    for i in range(1,bullet_len):
        desc_dict[bullet.cell(i,1).value]=f'{bullet.cell(i,2).value}\n'+f'{bullet.cell(i,3).value}\n'+f'{bullet.cell(i,4).value}\n'+f'{bullet.cell(i,5).value}\n'+f'{bullet.cell(i,6).value}\n'+f'{bullet.cell(i,7).value}\n'+f'{bullet.cell(i,8).value}\n'
    return desc_dict
def write_json():
    with open(os.getcwd()+ r"\package variables.json",'r+') as f:   #save packages list to JSON
        data=json.load(f)
        data['items_desc']= create_description_list()
        f.seek(0)
        json.dump(data,f,indent=4)   

def read_json():
    with open(os.getcwd()+ r"\package variables.json",'r+') as f:   #save packages list to JSON
        data=json.load(f)
        item_desc = data['items_desc']
        f.seek(0)
        json.dump(data,f,indent=4)
        return item_desc
        ## change to return entire json as dictionary
def write_descriptions(row, workbook, item_desc):
    desc = ''
    for i in range(3,10):
        sku = workbook.active.cell(row,i).value
        if sku!= None:
            desc +=str(item_desc[sku] + '\n')
    workbook.active.cell(row,11).value = desc  
