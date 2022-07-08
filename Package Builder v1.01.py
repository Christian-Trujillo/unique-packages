from tkinter.ttk import Button
import PySimpleGUI as gui
from types import NoneType
import openpyxl as xl
import os
from datetime import date
import json
import win32gui, win32con
import descriptions as de

### hides console window ###
hide = win32gui.GetForegroundWindow()
win32gui.ShowWindow(hide, win32con.SW_HIDE)

# setup package builder
packages = []
pkg = ["","","","","","",""]
categories={'RANGE':0,'COOKTOP':0,"RANGE HOOD":1,'DISHWASHER':2,'REFRIGERATOR':3,'WINE COOLER':4,'WALL OVEN':5,'MICROWAVE':6}

class Item():
    def __init__(self,sku, size,category_1):
        self.sku=sku
        try:
            self.size=int(size)
        except: self.size=size
        if self.size == 'AUX':
            self.size2 =[24,30,36,48,'AUX']
        else:
            try:
                self.size2 = [int(size),'AUX']
            except:self.size2 = [size,'AUX']
        if category_1 == 'RANGE':
            self.cat1="RANGE"
            self.cat2=['COOKTOP','WALL OVEN','RANGE']
        elif category_1 == 'WALL OVEN':
            self.cat1="WALL OVEN"
            self.cat2=['RANGE','WALL OVEN']
        elif category_1 == 'COOKTOP':
            self.cat1="COOKTOP"
            self.cat2=['COOKTOP','RANGE']
        elif category_1 == 'RANGE HOOD/ MICROWAVE':
            self.cat1="RANGE HOOD"
            self.cat2=['RANGE HOOD','MICROWAVE']
        else:
            self.cat1=category_1
            self.cat2=[category_1]

def prompt(string='yes'):
    if input(string)=='yes':
        return [],True
    else:
        with open('package variables.json','r+') as f:
            data=json.load(f) 
            return list(data['packages'].values()),False
def item_cat(list_of_lists, var =[]):
    with open('package variables.json','r+') as f:
            data=json.load(f) 
            for lst in list_of_lists:
                var2=[]
                for item in lst:
                    var2.append(data['items'][item]['category'])
                var.append(var2)
    return var, False
base_item_list = []
hood_list=[]
microwave_list=[]
refrigerator_list=[]
items=[]
skucat={}
n=0


#set up GUI
gui.theme('DarkAmber')
layout=[[gui.T("")],[gui.Button('Start Package Builder'),gui.Text('                  '),gui.Button('Update Descriptions')]]
window=gui.Window('Package Builder', layout, size=(800,200))


while True:
    event,values = window.read()

    #auto updates names of selected files
    if values['file1'] !='' or values['file2']!='':
        name_1 = values['file1']
        window['_file1_'].update(value=name_1)
        name_2 = values['file2']
        window['_file2_'].update(value=name_2)

    #close window
    if event == gui.WIN_CLOSED:
        break
    elif event == "Update Descriptions":
        de.write_json()
    #run builder by clicking submit
    elif event == 'Start Package Builder':
        if values['file1']=='':
            full_sku_path = os.getcwd() + r'\sku list & template\SKU List categorized.xlsx'
        else:
            full_sku_path = values['file1']
        if values['file2']=='':
            wkbk_path = os.getcwd() + r'\sku list & template\Packages template.xlsx'
        else:
            wkbk_path = values['file2']
        full_sku_wkbk = xl.load_workbook(full_sku_path)
        wkbk = xl.load_workbook(wkbk_path)
        full_sku= full_sku_wkbk['FULL SKU LIST']
        packages_wksht = wkbk.active
        data_len = full_sku.max_row +1

        for i in range(2,data_len):   #GET ITEMS INFO FROM SKU LIST, ADD TO ITEMS LIST
            exec(f'item_{i} = Item(full_sku.cell({i},1).value, full_sku.cell({i},3).value, full_sku.cell({i},2).value)')
            exec(f'items.append(item_{i})')
        for item in items:   # CREATE BASE ITEM LIST
            if item.cat1=="COOKTOP" :
                base_item_list.append(item)
            if item.cat1 == 'WALL OVEN':
                base_item_list.append(item)
            if item.cat1 == 'RANGE':
                base_item_list.append(item)
        for item in items:
            if item.cat1=="RANGE HOOD" :
                hood_list.append(item) 
        for item in items:
            if item.cat1=="REFRIGERATOR" :
                refrigerator_list.append(item) 
        for item in items:
            if item.cat1=="MICROWAVE" :
                microwave_list.append(item)
        
        #create 2s lists
        for item in base_item_list:
            for item2 in items:
                if item.cat1 == 'WALL OVEN' and item2.cat1 =='RANGEHOOD':
                    continue
                pack = ['','','','','','','']
                if (item.cat1 not in item2.cat2) and (item2.cat1 not in item.cat2) and (item.size in item2.size2) and type(item2.sku)!=NoneType:
                    pack[categories[item.cat1]] = item
                    pack[categories[item2.cat1]] = item2

                    if pack not in packages:
                        packages.append(pack)
                        n+=1
                        print(f'packages built: {n}', end="\r")
                        window['pckg'].update(value=str(n))
                        window.refresh()

        hood=True
        microwave=True
        refrigerator=True
        dishwasher=True 

        for pack in packages:
            m=0
            for item in pack:
                if item == '':
                    m+=1   #rest of list based on 2s
            if m==4 and hood:   #REMOVE ITEMS TO SHORTEN CHECK LIST, OPTIMIZING BUT NOT LIMITING PACKAGES
                for item in hood_list:
                        items.remove(item)
                        hood=False
            elif m==3 and microwave:
                for item in microwave_list:
                        items.remove(item)
                        microwave=False
            elif m==2 and refrigerator:
                for item in refrigerator_list:
                        items.remove(item)
                        refrigerator=False
            elif m==1 and dishwasher:
                for item in items:
                    if 'DISHWASHER' in item.cat2:
                        items.remove(item)
                        dishwaser=False
            if m==0:
                break 
            cats = []
            sizes = ['AUX']
            for sku in pack:  # list of categories and size in package
                if sku !='': 
                    for cat in sku.cat2:
                        cats.append(cat)
                    sizes.append(sku.size)
            for item in items:       
                if item.cat1 in cats :
                    continue
                if item.size not in sizes:
                    continue
                if item.cat1=='RANGEHOOD' and 'COOKTOP' not in cats:
                    continue
                if type(item.sku)!= NoneType:  
                    pckg=['','','','','','','']
                    for i in range(len(pack)):
                        pckg[i]=pack[i]
                    pckg[categories[item.cat1]] = item
                    if pckg not in packages:
                        packages.append(pckg)
                        n+=1
                        print(f'packages built: {n}', end="\r")
                        window['pckg'].update(value=str(n))
                        window.refresh()
        item_desc = de.read_json()
        for i in range(len (packages))  :
            serial = ''
            for j in range(len(packages[i])) :
                try:
                    packages_wksht.cell(i+2,j+3).value=packages[i][j].sku
                    serial+=packages[i][j].sku         
                except: pass
            packages_wksht.cell(i+2,10).value=serial
            de.write_descriptions(i+2,wkbk, item_desc)
            # packages_wksht.cell(i+2,10).value= str("hello")
        # with open(os.getcwd()+ r"\package variables.json",'r+') as f:   #save packages list to JSON
            #     data=json.load(f)
            #     for i in range(len(packages)):
            #         items=[]
            #         for item in packages[i]:
            #             items.append(item.sku)   
            #         data['packages'][f'pack_{i}']=items
            #     f.seek(0)
            #     json.dump(data,f,indent=4)   
        today = date.today()

        wkbk.save(f'full packages {today}.xlsx')
        window['pckg'].update(value=str(n)+'  ---  All Done!')
        window.refresh()

   
window.close()